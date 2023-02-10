#################################################################################
#
# This is a study on two libraries to update Excel files: openpyxl and xlwings
# This exercise has shown that openpyxl breaks the xlsx files in this repo (maybe
#   because of the macros, or the formulae), while xlwings works fine.
#
# This script reads a previously generated JSON file with the results of Azure
#   Resource Graph queries, and stores them in the 'Comments' column of a
#   spreadsheet. Both the JSON file and the spreadsheet file are supplied as
#   parameters.
# 
# Last updated: March 2022
#
#################################################################################

import json
import argparse
import sys
from pandas import DataFrame
from openpyxl import load_workbook
import xlwings as xw

# Get input arguments
parser = argparse.ArgumentParser(description='Update a checklist spreadsheet with JSON-formatted Azure Resource Graph results')
parser.add_argument('--graph-file', dest='graph_file', action='store',
                    help='You need to supply a JSON file containing the results of Azure Resource Graph Queries')
parser.add_argument('--excel-file', dest='excel_file', action='store',
                    help='You need to supply an Excel file where the query results will be stored')
parser.add_argument('--mode', dest='mode', action='store', default="openpyxl",
                    help='It can be either xlwings or openpyxl (default is openpyxl)')
parser.add_argument('--verbose', dest='verbose', action='store_true',
                    default=False,
                    help='run in verbose mode (default: False)')
args = parser.parse_args()
graph_file = args.graph_file
excel_file = args.excel_file
mode = args.mode

# Constants
guid_column_index = "K"
comment_column_index = "G"
sample_cell_index = 'A4'

# Get JSON
try:
    with open(graph_file) as f:
        graph_data = json.load(f)
except Exception as e:
    print("ERROR: Error when processing JSON file", graph_file, "-", str(e))
    sys.exit(1)

# Load workbook
try:
    if mode == 'openpyxl':
        if args.verbose:
            print("DEBUG: working with openpyxl library")
        wb = load_workbook(filename = excel_file)
        ws = wb['Checklist']
    elif mode == 'xlwings':
        if args.verbose:
            print("DEBUG: working with xlwings library")
        wb = xw.Book(excel_file)
        ws = wb.sheets['Checklist']
    else:
        print("ERROR: mode {0} not recognized".format(mode))
except Exception as e:
    print("ERROR: Error when opening Excel file", excel_file, "-", str(e))
    sys.exit(1)

# Print specific cell
if args.verbose:
    print("DEBUG: looking at spreadsheet for", ws[sample_cell_index].value)

# Get GUID column into a list
if mode == 'openpyxl':
    guid_col = ws[guid_column_index]
    guid_col_values = [x.value for x in guid_col]
    if args.verbose:
        print("DEBUG: GUID column retrieved with", str(len(guid_col_values)), "values")
elif mode == 'xlwings':
    guid_col_values = ws.range(guid_column_index + ":" + guid_column_index).value
    if args.verbose:
        print("DEBUG: GUID column retrieved with", str(len(guid_col_values)), "values")
else:
    print("ERROR: mode {0} not recognized".format(mode))
    sys.exit(1)

# Go over all checks in the JSON file
for check in graph_data['checks']:
    guid = check['guid']
    arm_id = check['id']
    compliant = check['compliant']
    if (compliant == "false"):
        comment = "Non-compliant: {0}\n".format(arm_id)
    elif (compliant == "true"):
        comment = "Compliant: {0}\n".format(arm_id)
    else:
        print("ERROR: compliant status {0} not recognized".format(compliant))
    # Find the guid in the list
    if guid in guid_col_values:
        row = guid_col_values.index(guid)
        cell_index = comment_column_index + str(row)
        print("DEBUG: updating cell", cell_index)
        if mode == 'openpyxl':
            ws[cell_index] = comment
        elif mode == 'xlwings':
            ws.range(cell_index).value = comment
    else:
        print("ERROR: could not find GUID {0} in the Excel list".format(guid))

# Saving file
if mode == 'openpyxl':
    print("DEBUG: saving workbook", excel_file)
    try:
        wb.save(excel_file)
    except Exception as e:
        print("ERROR: Error when saving Excel file", excel_file, "-", str(e))
        sys.exit(1)
elif mode == 'xlwings':
    print("DEBUG: saving workbook", excel_file)
    try:
        wb.save()
    except Exception as e:
        print("ERROR: Error when saving Excel file", excel_file, "-", str(e))
        sys.exit(1)
else:
    print("ERROR: mode {0} not recognized".format(mode))
