######################################################################
#
# This script combines all of the existing checklists into one big
#   checklist, and saves it in JSON and XLSX (macrofree) formats.
#
# Example usage:
# python3 ./scripts/create_master_checklist.py \
#   --input-folder="./checklists" \
#   --language="en" \
#   --excel-file="./spreadsheet/macrofree/review_checklist_master_empty.xlsx" \
#   --output-name="checklist.en.master" \
#   --json-output-folder="./checklists/" \
#   --xlsx-output-folder="./spreadsheet/macrofree/"
# 
# Last updated: March 2022
#
######################################################################

import json
import argparse
import sys
import os
import requests
import glob
import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter

# Get input arguments
parser = argparse.ArgumentParser(description='Update a checklist spreadsheet with JSON-formatted Azure Resource Graph results')
parser.add_argument('--input-folder', dest='input_folder', action='store',
                    help='Input folder where the checklists to merge are stored')
parser.add_argument('--language', dest='language', action='store', default='en',
                    help='if checklist files are specified, ignore the non-English ones and only generate a spreadsheet for the English version (default: False)')
parser.add_argument('--excel-file', dest='excel_file', action='store',
                    help='You need to supply an Excel file that will be taken as template to create the XLSX file with the checklist')
parser.add_argument('--json-output-folder', dest='json_output_folder', action='store',
                    help='Folder where to store the JSON output')
parser.add_argument('--xlsx-output-folder', dest='xlsx_output_folder', action='store',
                    help='Folder where to store the macro free Excel output')
parser.add_argument('--output-name', dest='output_name', action='store',
                    help='File name (without extension) for the output files (.json and .xlsx extensions will be added automatically)')
parser.add_argument('--add-services', dest='add_services', action='store_true',
                    default=False, help='If services field should be added to the checklist items (default: False)')
parser.add_argument('--no-excel', dest='no_excel', action='store_true',
                    default=False, help='If a macrofree Excel spreadsheet should not be generated')
parser.add_argument('--no-json', dest='no_json', action='store_true',
                    default=False, help='If a JSON file should not be generated')
parser.add_argument('--no-data-validation', dest='no_data_validation', action='store_true',
                    default=False, help='If data validation should be skipped when updating the Excel spreadsheet')
parser.add_argument('--no-links', dest='no_links', action='store_true',
                    default=False, help='If hyperlinks should be skipped when updating the Excel spreadsheet')
parser.add_argument('--stats', dest='stats', action='store_true',
                    default=False, help='If statistics about the newly generated checklist should be displayed')
parser.add_argument('--show-service', dest='show_service', action='store',
                    help='If you want to print on screen the checks corresponding to a given service (e.g. "VM", or "none")')
parser.add_argument('--print-random', dest='print_random', action='store', default=0, type=int,
                    help='Print a random list of items on screen (default is 0)')
parser.add_argument('--verbose', dest='verbose', action='store_true',
                    default=False,
                    help='run in verbose mode (default: False)')
args = parser.parse_args()

# Inspect a string and return a list of the services to which that string is related to
def get_services_from_string(input_string):
    service_dict = {
        "AppSvc": ["App Service", "webapp"],
        "ExpressRoute": ["ExpressRoute", "Gateway Subnet"],
        "VPN": ["VPN", "Point-to-Site", "Site-to-Site", "Gateway Subnet"],
        "FrontDoor": ["Front Door", "FrontDoor"],
        "AppGW": ["Application Gateway", "AppGW", "AGIC"],
        "SQL": ["SQL"],
        "AVD": ["AVD", "Virtual Desktop", "WVD", "MSIX"],
        "AKS": ["AKS", "Kubernetes"],
        "AVS": ["AVS", "Azure VMware Solution", "VMware"],
        "Firewall": ["Azure Firewall", "Firewall Manager"],
        "NVA": ["NVA", "Network Virtual Appliance"],
        "Bastion": ["Bastion"],
        "SAP": ["SAP"],
        "VM": ["VM ", "VM.", "VM'", "VMs", "Virtual Machine"],  # Characters in 'VMx' is to avoid matching 'VMware'
        "Storage": ["Storage", "Blob", "File", "Queue", "Table", "CORS"],
        "ACR": ["ACR", "Registry"],
        "AKV": ["AKV", "Key Vault", "Secrets", "Keys", "Certificates"],
        "ServiceBus": ["Service Bus", "ASB", "Queue", "Topic", "Relay"],
        "EventHubs": ["Event Hubs", "EventHubs", "Event Hub", "EH"],
        "CosmosDB": ["Cosmos DB", "CosmosDB"],
        "SAP": ["SAP"],
        "Sentinel": ["Sentinel"],
        "Entra": ["Entra", "AAD", "Azure AD", "Azure Active Directory", "PIM", "JIT", "Privileged Identity Management", "Just in Time", "Conditional Access", "MFA", "2FA", "Identity", "Identities", "B2B", "B2C"],
        "DDoS": ["DDoS", "Denial of Service"],
        "LoadBalancer": ["Load Balancer", "LB", "ILB", "SLB"],
        "DNS": ["DNS", "Domain Name System"],
        "TrafficManager": ["Traffic Manager", "TM"],
        "VNet": ["VNet", "Virtual Network", "NSG", "Network Security Group", "UDR", "User Defined Route", "IP Plan", "hub-and-spoke", "subnet"],
        "Defender": ["Defender", "Security Center"],
        "Subscriptions": ["Subscriptions", "Subscription", "Management Group"],
        "VWAN": ["VWAN", "Virtual WAN"],
        "ARS": ["ARS", "Route Server"],
        "Monitor": ["Monitor", "Log Analytics", "LogAnalytics", "Metrics", "Alerts"],
        "NetworkWatcher": ["Network Watcher", "NetworkWatcher", "Connection Monitor", "Flow logs"],
        "Arc": ["Arc ", "Arc-"],        # Otherwise it matches 'Architecture'
        "RBAC": ["RBAC", "role"],
        "Backup": ["Backup"],
        "ASR": ["ASR", "Site Recovery", "Disaster Recovery"],
        "AzurePolicy": ["Azure Policy", "Policy", "Policies"],
        "APIM": ["APIM", "API Management"],
        "AppProxy": ["App Proxy", "AppProxy"],
        "WAF": ["WAF", "Web Application Firewall"],
        "PrivateLink": ["Private Link", "PrivateLink", "Private Endpoint"],
        "Cost": ["Cost", "Budget"],
    }
    services = []
    for service in service_dict:
        for keyword in service_dict[service]:
            if keyword.lower() in input_string.lower():
                services.append(service)
    return list(set(services))

# Returns True if the checklist file is valid, False otherwise
# Used to skip certain checklists, such as the old

# Consolidate all checklists into one big checklist object
def get_consolidated_checklist(input_folder, language):
    # Initialize checklist object
    checklist_master_data = {
        'items': [],
        'metadata': {
            'name': 'Master checklist',
            'timestamp': datetime.date.today().strftime("%B %d, %Y")
        }
    }
    # Find all files in the input folder matching the pattern "language*.json"
    if args.verbose:
        print("DEBUG: looking for JSON files in folder", input_folder, "with pattern *.", language + ".json...")
    checklist_files = glob.glob(input_folder + "/*." + language + ".json")
    if args.verbose:
        print("DEBUG: found", len(checklist_files), "JSON files")
    for checklist_file in checklist_files:
        # Get JSON
        try:
            with open(checklist_file) as f:
                checklist_data = json.load(f)
                if args.verbose:
                    print("DEBUG: JSON file", checklist_file, "loaded successfully with {0} items".format(len(checklist_data["items"])))
                # Verify that the checklist is not deprecated
                if "metadata" in checklist_data and "state" in checklist_data["metadata"] and "deprecated" in checklist_data["metadata"]["state"].lower():
                    if args.verbose:
                        print("DEBUG: skipping deprecated checklist", checklist_file)
                else:
                    for item in checklist_data["items"]:
                        # Add field with the name of the checklist
                        item["checklist"] = checklist_data["metadata"]["name"]
                    # Add items to the master checklist
                    checklist_master_data['items'] += checklist_data['items']
                    # Replace the master checklist severities and status sections (for a given language they should be all the same)
                    checklist_master_data['severities'] = checklist_data['severities']
                    checklist_master_data['status'] = checklist_data['status']
        except Exception as e:
            print("ERROR: Error when processing JSON file", checklist_file, "-", str(e))
        # Optionally, browse the checklist items and add the services field
        if args.add_services:
            for item in checklist_master_data["items"]:
                # Get service from the checklist name
                services = []
                services += get_services_from_string(item["checklist"])
                services += get_services_from_string(item["text"])
                services += get_services_from_string(item["category"])
                services += get_services_from_string(item["subcategory"])
                if "description" in item:
                    services += get_services_from_string(item["description"])
                item["services"] = list(set(services))
    if args.verbose:
        print("DEBUG: master checklist contains", len(checklist_master_data["items"]), "items")
    return checklist_master_data

# Print statistics about the checklist
def print_stats(checklist):
    print("INFO: Number of checks:", len(checklist["items"]))
    print("INFO: Number of categories:", len(set([item["category"] for item in checklist["items"]])))
    print("INFO: Number of items with no GUID:", len([item for item in checklist["items"] if "guid" not in item]))
    if args.add_services:
        print("INFO: Number of services:", len(set([service for item in checklist["items"] for service in item["services"]])))
        print("INFO: Number of items with no services:", len([item for item in checklist["items"] if len(item["services"]) == 0]))
        items = []
        if args.show_service:
            if args.verbose:
                print ("DEBUG: Getting items for service", args.show_service, "...")
            if args.show_service == "none":
                items = [item for item in checklist["items"] if len(item["services"]) == 0]
            else:
                items = [item for item in checklist["items"] if args.show_service.lower() in [x.lower() for x in item["services"]]]
        for item in items:
            print_item(item)

# Print on screen a given checklist item in a single line with fixed field widths
def print_item(item):
    text= item["text"]
    cat = item["category"]
    subcat = item["subcategory"]
    id = item["id"] if "id" in item else ""
    checklist = item["checklist"] if "checklist" in item else ""
    svcs = str(item["services"]) if "services" in item else ""
    print("{0: <25.25} {1: <10.10} {2: <25.25} {3: <25.25} {4: <80.80} {5: <25.25}".format(checklist, id, cat, subcat, text, svcs))

# Dump JSON object to file
def dump_json_file(json_object, filename):
    if args.verbose:
        print("DEBUG: dumping JSON object to file", filename)
    json_string = json.dumps(json_object, sort_keys=True, ensure_ascii=False, indent=4, separators=(',', ': '))
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(json_string)
        f.close()

# Format string so that it is compatible with Excel
def format4excel(input_string):
    # Remove equals sign at the beginning of the string
    if input_string and input_string[0] == "=":
        input_string = input_string[1:]
    # Return formatted string
    return input_string

# Create macro-free Excel file with the checklist
def update_excel_file(input_excel_file, output_excel_file, checklist_data):
    # Constants
    worksheet_checklist_name = 'Checklist'
    row1 = 8        # First row after which the Excel spreadsheet will be updated
    col_checklist_name = "A"
    row_checklist_name = "4"
    guid_column_index = "M"
    comment_column_index = "I"
    sample_cell_index = 'A4'
    col_checklist="A"
    col_area = "B"
    col_subarea = "C"
    col_waf_pillar = "D"
    col_services = "E"
    col_check = "F"
    col_desc = "G"
    col_sev = "H"
    col_status = "I"
    col_comment = "J"
    col_link = "K"
    col_training = "L"
    col_arg = "M"
    col_guid = "N"
    info_link_text = 'More info'
    training_link_text = 'Training'
    worksheet_values_name = 'Values'
    values_row1 = 2
    col_values_severity = "A"
    col_values_status = "B"
    col_values_area = "C"
    col_values_description = "H"
    last_column = col_guid

    # Load workbook
    try:
        wb = load_workbook(filename = input_excel_file)
        if args.verbose:
            print("DEBUG: workbook", input_excel_file, "opened successfully")
    except Exception as e:
        print("ERROR: Error when opening Excel file", input_excel_file, "-", str(e))
        sys.exit(1)

    # Get worksheet
    try:
        ws = wb[worksheet_checklist_name]
        if args.verbose:
            print("DEBUG: worksheet", worksheet_checklist_name, "selected successfully")
    except Exception as e:
        print("ERROR: Error when selecting worksheet", worksheet_checklist_name, "-", str(e))
        sys.exit(1)

    # Set checklist name
    try:
        ws[col_checklist_name + row_checklist_name] = checklist_data["metadata"]["name"]
        if args.verbose:
            print("DEBUG: starting filling the Excel spreadsheet with the values of checklist '{0}'".format(checklist_data["metadata"]["name"]))
    except Exception as e:
        print("ERROR: Error when selecting worksheet", worksheet_checklist_name, "-", str(e))
        sys.exit(1)

    # Get default status from the JSON, default to "Not verified"
    try:
        status_list = checklist_data.get("status")
        default_status = status_list[0].get("name")
        if args.verbose:
            print ("DEBUG: default status retrieved from checklist: '{0}'".format(default_status))
    except:
        default_status = "Not verified"
        if args.verbose:
            print ("DEBUG: Using default status 'Not verified'")
        pass

    # For each checklist item, add a row to spreadsheet
    row_counter = row1
    for item in checklist_data.get("items"):
        # Read variables from JSON
        checklist_name = format4excel(item.get("checklist"))
        guid = format4excel(item.get("guid"))
        category = format4excel(item.get("category"))
        subcategory = format4excel(item.get("subcategory"))
        waf_pillar = format4excel(item.get("waf"))
        text = format4excel(item.get("text"))
        description = format4excel(item.get("description"))
        severity = format4excel(item.get("severity"))
        link = format4excel(item.get("link"))
        training = format4excel(item.get("training"))
        status = default_status
        graph_query = format4excel(item.get("graph"))
        # Transform services array in a comma-separated string
        services = ""
        if "services" in item:
            for service in item["services"]:
                if len(services) > 0:
                    services += ", "
                services += service
        # Update Excel
        ws[col_checklist + str(row_counter)].value = checklist_name
        ws[col_area + str(row_counter)].value = category
        ws[col_subarea + str(row_counter)].value = subcategory
        ws[col_waf_pillar + str(row_counter)].value = waf_pillar
        ws[col_services + str(row_counter)].value = services
        ws[col_check + str(row_counter)].value = text
        ws[col_desc + str(row_counter)].value = description
        ws[col_sev + str(row_counter)].value = severity
        ws[col_status + str(row_counter)].value = status
        if not args.no_links:
            if link:
                ws[col_link + str(row_counter)].value = info_link_text
                ws[col_link + str(row_counter)].hyperlink = link
                ws[col_link + str(row_counter)].style = "Hyperlink"
                # ws[col_link + str(row_counter)].value = '=HYPERLINK("{}", "{}")'.format(link, info_link_text)
            if training:
                ws[col_training + str(row_counter)].value = training_link_text
                ws[col_training + str(row_counter)].value = training
                ws[col_training + str(row_counter)].style = "Hyperlink"
                # ws[col_training + str(row_counter)].value = '=HYPERLINK("{}", "{}")'.format(training, training_link_text)
        ws[col_arg + str(row_counter)].value = graph_query
        ws[col_guid + str(row_counter)].value = guid
        # Next row
        row_counter += 1

    # Create table
    # Corrupts file!!!!
    # table_ref = "A" + str(row1 - 1) + ":" + last_column + str(row_counter - 1)
    # if args.verbose:
    #     print("DEBUG: creating table for range {0}...".format(table_ref))
    # table = Table(displayName="Checklist", ref=table_ref)
    # ws.add_table(table)

    # Get number of checks
    number_of_checks = row_counter - row1
    
    # Display summary
    if args.verbose:
        print("DEBUG:", str(number_of_checks), "checks added to Excel spreadsheet")

    # Get worksheet
    try:
        wsv = wb[worksheet_values_name]
        if args.verbose:
            print("DEBUG: worksheet", worksheet_values_name, "selected successfully")
    except Exception as e:
        print("ERROR: Error when selecting worksheet", worksheet_values_name, "-", str(e))
        sys.exit(1)

    # Update status
    row_counter = values_row1
    for item in checklist_data.get("status"):
        status = item.get("name")
        description = item.get("description")
        wsv[col_values_status + str(row_counter)].value = status
        wsv[col_values_description + str(row_counter)].value = description
        row_counter += 1

    # Display summary
    if args.verbose:
        print("DEBUG:", str(row_counter - values_row1), "statuses addedd to Excel spreadsheet")

    # Update severities
    row_counter = values_row1
    for item in checklist_data.get("severities"):
        severity = item.get("name")
        wsv[col_values_severity + str(row_counter)].value = severity
        row_counter += 1

    # Display summary
    if args.verbose:
        print("DEBUG:", str(row_counter - values_row1), "severities addedd to Excel spreadsheet")

    # Data validation
    # UserWarning: Data Validation extension is not supported and will be removed!!!!
    # dv = DataValidation(type="list", formula1='=Values!$B$2:$B$6', allow_blank=True, showDropDown=True)
    if not args.no_data_validation:
        dv = DataValidation(type="list", formula1='=Values!$B$2:$B$6', allow_blank=True)
        rangevar = col_status + str(row1) +':' + col_status + str(row1 + number_of_checks)
        if args.verbose:
            print("DEBUG: adding data validation to range", rangevar)
        dv.add(rangevar)
        ws.add_data_validation(dv)

    # Close book
    if args.verbose:
        print("DEBUG: saving workbook", output_excel_file)
    try:
        wb.save(output_excel_file)
    except Exception as e:
        print("ERROR: Error when saving Excel file to", output_excel_file, "-", str(e))
        sys.exit(1)

########
# Main #
########

# Download checklist
if args.input_folder:
    # Get consolidated checklist
    checklist_master_data = get_consolidated_checklist(args.input_folder, args.language)
    # Dump master checklist to JSON file
    if not args.no_json:
        json_output_file = os.path.join(args.json_output_folder, args.output_name + ".json")
        dump_json_file(checklist_master_data, json_output_file)
    # Update spreadsheet
    if not args.no_excel:
        xlsx_output_file = os.path.join(args.xlsx_output_folder, args.output_name + ".xlsx")
        update_excel_file(args.excel_file, xlsx_output_file, checklist_master_data)
    # Print random items
    if args.print_random > 0:
        import random
        random_items = random.sample(checklist_master_data["items"], int(args.print_random))
        for item in random_items:
            print_item(item)
    # Show statistics
    if args.stats:
        print_stats(checklist_master_data)
else:
    print("ERROR: No input folder specified")
