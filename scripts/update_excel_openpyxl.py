######################################################################
#
# This script reads the checklist items from the latest checklist file
#   in Github (or from a local file) and populates an Excel spreadsheet
#   with the contents.
# 
# Last updated: March 2022
#
######################################################################

import json
import argparse
import sys
import os
import requests
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Get input arguments
parser = argparse.ArgumentParser(description='Update a checklist spreadsheet with JSON-formated Azure Resource Graph results')
parser.add_argument('--checklist-file', dest='checklist_file', action='store',
                    help='You can optionally supply a JSON file containing the checklist you want to dump to the Excel spreadsheet. Otherwise it will take the latest file from Github')
parser.add_argument('--only-english', dest='only_english', action='store_true', default=False,
                    help='if checklist files are specified, ignore the non-English ones and only generate a spreadsheet for the English version (default: False)')
parser.add_argument('--technology', dest='technology', action='store',
                    help='If you do not supply a JSON file with the checklist, you need to specify the technology from which the latest checklist will be downloaded from Github')
parser.add_argument('--excel-file', dest='excel_file', action='store',
                    help='You need to supply an Excel file where the checklist will be written')
parser.add_argument('--output-excel-file', dest='output_excel_file', action='store',
                    help='You can optionally supply an Excel file where the checklist will be saved, otherwise it will be updated in-place')
parser.add_argument('--output-path', dest='output_path', action='store',
                    help='If using --output-name-is-input-name, folder where to store the results')
parser.add_argument('--output-name-is-input-name', dest='output_name_is_input_name', action='store_true',
                    default=False,
                    help='Save the output in a file with the same filename as the JSON input, but with xlsx extension')
parser.add_argument('--verbose', dest='verbose', action='store_true',
                    default=False,
                    help='run in verbose mode (default: False)')
args = parser.parse_args()
checklist_file = args.checklist_file
excel_file = args.excel_file
technology = args.technology

# Constants
worksheet_checklist_name = 'Checklist'
row1 = 8        # First row after which the Excel spreadsheet will be updated
col_checklist_name = "A"
row_checklist_name = "4"
guid_column_index = "L"
comment_column_index = "G"
sample_cell_index = 'A4'
col_area = "A"
col_subarea = "B"
col_check = "C"
col_desc = "D"
col_sev = "E"
col_status = "F"
col_comment = "G"
col_link = "H"
col_training = "I"
col_arg_success = "J"
col_arg_failure = "K"
col_guid = "L"
info_link_text = 'More info'
training_link_text = 'Training'
worksheet_values_name = 'Values'
values_row1 = 2
col_values_severity = "A"
col_values_status = "B"
col_values_area = "C"
col_values_description = "H"

# Main function
def update_excel_file(input_excel_file, output_excel_file, checklist_data):
    # Load workbook
    try:
        wb = load_workbook(filename = input_excel_file)
        if args.verbose:
            print("DEBUG: workbook", input_excel_file, "opened successfully")
    except Exception as e:
        print("ERROR: Error when opening Excel file", input_excel_file, "-", str(e))
        sys.exit(1)

    # Get worksheet
    try:
        ws = wb[worksheet_checklist_name]
        if args.verbose:
            print("DEBUG: worksheet", worksheet_checklist_name, "selected successfully")
    except Exception as e:
        print("ERROR: Error when selecting worksheet", worksheet_checklist_name, "-", str(e))
        sys.exit(1)

    # Set checklist name
    try:
        ws[col_checklist_name + row_checklist_name] = checklist_data["metadata"]["name"]
        if args.verbose:
            print("DEBUG: starting filling the Excel spreadsheet with the values of checklist '{0}'".format(checklist_data["metadata"]["name"]))
    except Exception as e:
        print("ERROR: Error when selecting worksheet", worksheet_checklist_name, "-", str(e))
        sys.exit(1)

    # Get default status from the JSON, default to "Not verified"
    try:
        status_list = checklist_data.get("status")
        default_status = status_list[0].get("name")
        if args.verbose:
            print ("DEBUG: default status retrieved from checklist: '{0}'".format(default_status))
    except:
        default_status = "Not verified"
        if args.verbose:
            print ("DEBUG: Using default status 'Not verified'")
        pass

    # For each checklist item, add a row to spreadsheet
    row_counter = row1
    for item in checklist_data.get("items"):
        # Read variables from JSON
        guid = item.get("guid")
        category = item.get("category")
        subcategory = item.get("subcategory")
        text = item.get("text")
        description = item.get("description")
        severity = item.get("severity")
        link = item.get("link")
        training = item.get("training")
        status = default_status
        graph_query_success = item.get("graph_success")
        graph_query_failure = item.get("graph_failure")
        # Update Excel
        ws[col_area + str(row_counter)].value = category
        ws[col_subarea + str(row_counter)].value = subcategory
        ws[col_check + str(row_counter)].value = text
        ws[col_desc + str(row_counter)].value = description
        ws[col_sev + str(row_counter)].value = severity
        ws[col_status + str(row_counter)].value = status
        ws[col_link + str(row_counter)].value = link
        # if link != None:
        #     link_elements = link.split('#')
        #     link_address = link_elements[0]
        #     if len(link_elements) > 1:
        #         link_subaddress = link_elements[1]
        #     else:
        #         link_subaddress = ""
        #     ws.api.Hyperlinks.Add (Anchor=ws[col_link + str(row_counter)].api, Address=link_address, SubAddress=link_subaddress, ScreenTip="", TextToDisplay=info_link_text)
        ws[col_training + str(row_counter)].value = training
        # if training != None:
        #     training_elements = training.split('#')
        #     training_address = training_elements[0]
        #     if len(training_elements) > 1:
        #         training_subaddress = training_elements[1]
        #     else:
        #         training_subaddress = ""
        #     ws.api.Hyperlinks.Add (Anchor=ws[col_training + str(row_counter)].api, Address=training_address, SubAddress=training_subaddress, ScreenTip="", TextToDisplay=training_link_text)
        # GUID and ARG queries
        ws[col_arg_success + str(row_counter)].value = graph_query_success
        ws[col_arg_failure + str(row_counter)].value = graph_query_failure
        ws[col_guid + str(row_counter)].value = guid
        # Next row
        row_counter += 1

    # Display summary
    if args.verbose:
        number_of_checks = row_counter - row1
        print("DEBUG:", str(number_of_checks), "checks addedd to Excel spreadsheet")

    # Get worksheet
    try:
        wsv = wb[worksheet_values_name]
        if args.verbose:
            print("DEBUG: worksheet", worksheet_values_name, "selected successfully")
    except Exception as e:
        print("ERROR: Error when selecting worksheet", worksheet_values_name, "-", str(e))
        sys.exit(1)

    # Update categories
    row_counter = values_row1
    for item in checklist_data.get("categories"):
        area = item.get("name")
        wsv[col_values_area + str(row_counter)].value = area
        row_counter += 1

    # Display summary
    if args.verbose:
        print("DEBUG:", str(row_counter - values_row1), "categories addedd to Excel spreadsheet")

    # Update status
    row_counter = values_row1
    for item in checklist_data.get("status"):
        status = item.get("name")
        description = item.get("description")
        wsv[col_values_status + str(row_counter)].value = status
        wsv[col_values_description + str(row_counter)].value = description
        row_counter += 1

    # Display summary
    if args.verbose:
        print("DEBUG:", str(row_counter - values_row1), "statuses addedd to Excel spreadsheet")

    # Update severities
    row_counter = values_row1
    for item in checklist_data.get("severities"):
        severity = item.get("name")
        wsv[col_values_severity + str(row_counter)].value = severity
        row_counter += 1

    # Display summary
    if args.verbose:
        print("DEBUG:", str(row_counter - values_row1), "severities addedd to Excel spreadsheet")

    # Data validation
    # dv = DataValidation(type="list", formula1='=Values!$B$2:$B$6', allow_blank=True, showDropDown=True)
    dv = DataValidation(type="list", formula1='=Values!$B$2:$B$6', allow_blank=True)
    rangevar = col_status + str(row1) +':' + col_status + str(row1 + number_of_checks)
    if args.verbose:
        print("DEBUG: adding data validation to range", rangevar)
    dv.add(rangevar)
    ws.add_data_validation(dv)

    # Close book
    if args.verbose:
        print("DEBUG: saving workbook", output_excel_file)
    try:
        wb.save(output_excel_file)
    except Exception as e:
        print("ERROR: Error when saving Excel file to", output_excel_file, "-", str(e))
        sys.exit(1)

########
# Main #
########

# Download checklist
if checklist_file:
    checklist_file_list = checklist_file.split(" ")
    # If only-english parameter was supplied, take only the English version and remove duplicates
    if args.only_english:
        checklist_file_list = [file[:-8] + '.en.json' for file in checklist_file_list]
        checklist_file_list = list(set(checklist_file_list))
    for checklist_file in checklist_file_list:
        if args.verbose:
            print("DEBUG: Opening checklist file", checklist_file)
        # Get JSON
        try:
            with open(checklist_file) as f:
                checklist_data = json.load(f)
        except Exception as e:
            print("ERROR: Error when processing JSON file", checklist_file, "-", str(e))
            sys.exit(0)
        # Set input and output files
        input_excel_file = excel_file
        if args.output_excel_file:
            output_excel_file = args.output_excel_file
        elif args.output_name_is_input_name:
            if args.output_path:
                # Get filename without path and extension
                output_excel_file = os.path.splitext(os.path.basename(checklist_file))[0] + '.xlsx'
                output_excel_file = os.path.join(args.output_path, output_excel_file)
            else:
                # Just change the extension
                output_excel_file = os.path.splitext(checklist_file)[0] + '.xlsx'
        # Update spreadsheet
        update_excel_file(input_excel_file, output_excel_file, checklist_data)
else:
    if technology:
        checklist_url = "https://raw.githubusercontent.com/Azure/review-checklists/main/checklists/" + technology + "_checklist.en.json"
    else:
        checklist_url = "https://raw.githubusercontent.com/Azure/review-checklists/main/checklists/lz_checklist.en.json"
    if args.verbose:
        print("DEBUG: Downloading checklist file from", checklist_url)
    response = requests.get(checklist_url)
    # If download was successful
    if response.status_code == 200:
        if args.verbose:
            print ("DEBUG: File {0} downloaded successfully".format(checklist_url))
        try:
            # Deserialize JSON to object variable
            checklist_data = json.loads(response.text)
        except Exception as e:
            print("Error deserializing JSON content: {0}".format(str(e)))
            sys.exit(1)
    # Upload spreadsheet
    if args.output_excel_file:
        output_excel_file = args.output_excel_file
    else:
        output_excel_file = excel_file
    update_excel_file(excel_file, output_excel_file, checklist_data)

